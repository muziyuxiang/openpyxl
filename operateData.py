# coded by ligang(2020-02-13)
# 注意事项：两个表格格式需和示例表格相同
# 使用说明：需要修改5个数据：(1)num_countries(2)num_data(3)三个路径 分别为数据文件路径、国家名称文件路径、输出结果路径
import numpy as np
import pandas as pd
from openpyxl import load_workbook

num_countries = 142  # 行数 真实值+1
num_data = 2022  # 行数 真实值+1
wb1 = load_workbook("/Users/ligang/Downloads/数据.xlsx")  # 数据文件路径
ws1 = wb1.active
rows1 = []
for row1 in ws1.iter_rows():
    rows1.append(row1)  # 读出excel中所有的列
wb2 = load_workbook("/Users/ligang/Downloads/数据2.xlsx")  # 国家文件路径
ws2 = wb2.active
cols2 = []
for col2 in ws2.iter_cols():
    cols2.append(col2)

countries = ['' for i in range(num_countries)]  # 存储国家的字符串数组
for i in range(1, num_countries):
    # print(cols2[0][i].value)
    countries[i] = cols2[0][i].value
    # print(countries[i])


# 计算每一个国家的出口公司个数
def calculate_percountry():
    count = np.zeros([num_countries, 1], dtype=np.int)  # count为每个国家的出口总数
    for i in range(1, num_countries):
        # print(cols2[0][i].value)
        for j in range(1, num_data):
            if rows1[j][1].value == cols2[0][i].value:
                count[i] += 1
    return count


cou = calculate_percountry()
# print(cou)


# 计算每两个国家共同的出口公司个数
def calculate_intercountry():
    # 建立字典 表示对每个国家出口的所有公司
    dict = {}
    for country in countries:
        list_company = []
        for i in range(1, num_data):
            if rows1[i][1].value == country:
                list_company.append(rows1[i][0].value)  # 把对同一个国家出口的所有公司加进列表
                # print(country, rows1[i][0].value)
        dict[country] = list_company

    count = np.zeros([num_countries, num_countries], dtype=np.int) # 计数矩阵
    for i in range(1, num_countries):
        for j in range(1, num_countries):
            temp_count = 0
            country1 = countries[i]
            country2 = countries[j]
            list1 = dict[country1]  # 每个国家分别所对应的出口公司
            list2 = dict[country2]
            for company in list1:
                if list2.__contains__(company):
                    temp_count += 1
                    # print(i,j, temp_count)
            count[i][j] = temp_count
    return count

def calculate_final():
    per_count = calculate_percountry()
    inter_count = calculate_intercountry()
    final_result = np.zeros([num_countries, num_countries], dtype=np.float)
    for i in range(1, num_countries):
        for j in range(1, num_countries):
            temp = max(per_count[i], per_count[j])
            if temp == 0:
                final_result[i][j] = -1
            else:
                final_result[i][j] = inter_count[i][j] / temp
    return final_result


final_result = calculate_final()
final_result = final_result[1:,1:]
data = pd.DataFrame(final_result)
writer = pd.ExcelWriter("/Users/ligang/Downloads/输出结果2.xlsx")
data.to_excel(writer, 'page_1', float_format="%.5f")
writer.save()
writer.close()
wb1.close()
wb2.close()

