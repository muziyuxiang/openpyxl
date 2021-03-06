# coded by ligang(2020-02-19)
import numpy as np
import pandas as pd
import openpyxl

num_countries = 142  # 行数 真实值+1
num_part_countries = 75  # 行数 真实值+1

wb1 = openpyxl.load_workbook("/Users/ligang/Downloads/2.1出口-城市层面-出口市场.xlsx")  # 数据文件路径
ws1 = wb1["张掖"]  # 获取excel中的一个表单
cols1 = []
for col1 in ws1.iter_cols():
    cols1.append(col1)  # 获取每个表单中的每一列

wb2 = openpyxl.load_workbook("/Users/ligang/Downloads/市场相似性.xlsx")  # 数据文件路径
ws2 = wb2.active
cols2 = []
for col2 in ws2.iter_cols():
    cols2.append(col2)  # 获取每个表单中的每一列

countries = ['' for i in range(num_countries)]  # 存储国家的字符串数组
for i in range(1, num_countries):
    # print(cols2[0][i].value)
    countries[i] = cols1[0][i].value
    # print(countries[i])

part_countries = ['' for i in range(num_part_countries)]  # 存储国家的字符串数组
for i in range(1, num_part_countries):
    # print(cols2[0][i].value)
    part_countries[i] = cols1[1][i].value

similar_countries = ['' for i in range(num_countries)]  # 存储国家的字符串数组
for i in range(1, num_countries):
    similar_countries[i] = cols2[0][i].value
    # print(similar_countries[i])

cols = np.zeros(num_part_countries, dtype=np.int)   # 存储列数
for i in range(1, num_part_countries):
    cols[i] = similar_countries.index(part_countries[i])

values = np.zeros(num_countries)
for i in range(1, num_countries):
    row = similar_countries.index(countries[i])
    fenzi = 0
    fenmu = 0
    for j in range(1, num_countries):
        fenmu += cols2[j][row].value
    for j in range(1, num_part_countries):
        fenzi += cols2[cols[j]][row].value
    values[i] = fenzi*1.0/fenmu


data = pd.DataFrame(values)
writer = pd.ExcelWriter("/Users/ligang/Downloads/输出结果.xlsx")
data.to_excel(writer, 'page_1', float_format="%.5f")
writer.save()
writer.close()
wb1.close()
wb2.close()
